#!/usr/bin/env python3
"""Generate lookup JSON files from xlsx and existing universities.json."""

from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl required: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[2]
LOOKUPS = ROOT / "data" / "lookups"
XLSX = ROOT / "assets" / "LPDP-Beasiswa" / "LPDP 2026.xlsx"
JSON_PATH = ROOT / "frontend" / "universities.json"

BIDANG_ALIASES = {
    "Hilirisasi & Industrialisasi": "Hilirisasi",
    "Hilirisasi dan Industrialisasi": "Hilirisasi",
    "Material dan Manufaktur": "Material dan Manufaktur",
    "Kebijakan Publik & Hukum": "Kebijakan Publik dan Hukum",
    "Kebijakan Publik dan Hukum": "Kebijakan Publik dan Hukum",
}

GARBLE_ALIASES = {
    "IEtaITlia4SEMM": "Italia",
    "SKc)anada": "Canada",
    "FItAalSia)": "Italia",
    ")Tiongkok": "Tiongkok",
    "uTriinogngkok": "Tiongkok",
    ")Jepang": "Jepang",
}


def clean_cell(val) -> str:
    if val is None:
        return ""
    return re.sub(r"\s+", " ", str(val).replace("\n", " ")).strip()


def build_country_region() -> dict:
    """Map PDF country names -> {lokasi, region}."""
    mapping: dict[str, dict[str, str]] = {}

    if JSON_PATH.exists():
        with open(JSON_PATH, encoding="utf-8") as f:
            data = json.load(f)
        for row in data:
            if row.get("Tipe") != "Luar Negeri":
                continue
            lokasi = clean_cell(row.get("Lokasi"))
            region = clean_cell(row.get("Region"))
            if lokasi:
                mapping[lokasi] = {"lokasi": lokasi, "region": region}
                mapping[lokasi.lower()] = {"lokasi": lokasi, "region": region}

    if XLSX.exists():
        wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
        ws = wb["luar negeri"]
        for r in ws.iter_rows(min_row=2, values_only=True):
            negara = clean_cell(r[6])
            benua = clean_cell(r[7])
            if negara:
                entry = {"lokasi": negara, "region": benua}
                mapping[negara] = entry
                mapping[negara.lower()] = entry
        wb.close()

    # Common PDF variants (English/local names)
    extras = {
        "Denmark": ("Denmark", "Eropa"),
        "Italia": ("Italia", "Eropa"),
        "Germany": ("Jerman", "Eropa"),
        "United Kingdom": ("Inggris", "Eropa"),
        "United States": ("Amerika Serikat", "Amerika"),
        "USA": ("Amerika Serikat", "Amerika"),
        "UK": ("Inggris", "Eropa"),
        "China": ("Tiongkok", "Asia"),
        "South Korea": ("Korea Selatan", "Asia"),
        "New Zealand": ("Selandia Baru", "Oseania"),
        "Netherlands": ("Belanda", "Eropa"),
        "France": ("Prancis", "Eropa"),
        "Japan": ("Jepang", "Asia"),
        "Australia": ("Australia", "Oseania"),
        "Canada": ("Kanada", "Amerika"),
        "Republik Rakyat Tiongkok": ("Tiongkok", "Asia"),
        "RRT": ("Tiongkok", "Asia"),
        "Oman": ("Oman", "Asia"),
        "Singapore": ("Singapura", "Asia"),
        "Hong Kong": ("Hong Kong", "Asia"),
        "Taiwan": ("Taiwan", "Asia"),
        "Sweden": ("Swedia", "Eropa"),
        "Switzerland": ("Swiss", "Eropa"),
        "Finland": ("Finlandia", "Eropa"),
        "Norway": ("Norwegia", "Eropa"),
        "Belgium": ("Belgia", "Eropa"),
        "Austria": ("Austria", "Eropa"),
        "Ireland": ("Irlandia", "Eropa"),
        "Spain": ("Spanyol", "Eropa"),
        "Russia": ("Rusia", "Eropa"),
        "India": ("India", "Asia"),
        "Malaysia": ("Malaysia", "Asia"),
        "Thailand": ("Thailand", "Asia"),
        "Saudi Arabia": ("Arab Saudi", "Asia"),
    }
    for src, (lokasi, region) in extras.items():
        entry = {"lokasi": lokasi, "region": region}
        mapping[src] = entry
        mapping[src.lower()] = entry

    # Canonical list keyed by lokasi
    canonical: dict[str, dict[str, str]] = {}
    for v in mapping.values():
        canonical[v["lokasi"]] = v
    return canonical


def build_pt_province() -> dict:
    """Map university name -> {lokasi: provinsi, region: pulau}."""
    mapping: dict[str, dict[str, str]] = {}

    sources = []
    if XLSX.exists():
        wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
        for sheet in ("dalam negeri", "afirmasi"):
            if sheet in wb.sheetnames:
                sources.append(wb[sheet])
    else:
        sources = []

    for ws in sources:
        for r in ws.iter_rows(min_row=2, values_only=True):
            pt = clean_cell(r[4])
            prov = clean_cell(r[6])
            pulau = clean_cell(r[7])
            if pt and prov:
                mapping[pt] = {"lokasi": prov, "region": pulau}

    if JSON_PATH.exists():
        with open(JSON_PATH, encoding="utf-8") as f:
            data = json.load(f)
        for row in data:
            if row.get("Tipe") != "Dalam Negeri":
                continue
            pt = clean_cell(row.get("Perguruan Tinggi"))
            lokasi = clean_cell(row.get("Lokasi"))
            region = clean_cell(row.get("Region"))
            if pt and lokasi and pt not in mapping:
                mapping[pt] = {"lokasi": lokasi, "region": region}

    return mapping


def build_bidang_aliases() -> dict:
    aliases = dict(BIDANG_ALIASES)
    if XLSX.exists():
        wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
        ws = wb["luar negeri"]
        raw_to_norm: dict[str, str] = defaultdict(set)
        for r in ws.iter_rows(min_row=2, values_only=True):
            norm = clean_cell(r[3])
            if norm:
                raw_to_norm[norm].add(norm)
        wb.close()
    return aliases


def main() -> int:
    LOOKUPS.mkdir(parents=True, exist_ok=True)

    country = build_country_region()
    pt = build_pt_province()
    bidang = build_bidang_aliases()
    garble = dict(GARBLE_ALIASES)

    (LOOKUPS / "country_region.json").write_text(
        json.dumps(country, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (LOOKUPS / "pt_province.json").write_text(
        json.dumps(pt, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (LOOKUPS / "bidang_aliases.json").write_text(
        json.dumps(bidang, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (LOOKUPS / "garble_aliases.json").write_text(
        json.dumps(garble, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(f"country_region: {len(country)} entries")
    print(f"pt_province: {len(pt)} entries")
    print(f"bidang_aliases: {len(bidang)} entries")
    print(f"garble_aliases: {len(garble)} entries")
    return 0


if __name__ == "__main__":
    sys.exit(main())
